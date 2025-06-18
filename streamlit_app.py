import streamlit as st
import pandas as pd
import openpyxl
import unicodedata
import re
from io import BytesIO
from zipfile import ZipFile
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# === UTILS ===
def remove_accents(s):
    if isinstance(s, str):
        return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s

def find_column(df, targets):
    for col in df.columns:
        col_norm = remove_accents(str(col)).lower().replace(' ', '').replace('_','')
        for t in targets:
            t_norm = remove_accents(t).lower().replace(' ', '').replace('_','')
            if col_norm.startswith(t_norm) or t_norm in col_norm:
                return col
    return None

def format_phone_number(phone):
    if isinstance(phone, str) and phone.replace(".", "").isdigit():
        phone = phone.split(".")[0]
        phone = phone.zfill(10)
        return " ".join([phone[i:i+2] for i in range(0, len(phone), 2)])
    return phone

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# === OPT PROCESSING ===
def format_name_advanced(value):
    if isinstance(value, str):
        value = remove_accents(value)
        return " ".join(word.capitalize() for word in value.split())
    return value

def process_opt(df, ref_bytes):
    df.columns = [remove_accents(str(col)) for col in df.columns]
    mapping = {
        "Civilite": ["Civilite", "CivilitE", "Civ", "Civilité"],
        "Nom": ["Nom"],
        "Prenom": ["Prenom", "PrEnom", "Prénom"],
        "Adresse": ["Adresse"],
        "Code Postal": ["Code Postal", "CP"],
        "Ville": ["Ville"],
        "Pays": ["Pays"],
        "Email": ["Email", "Mail", "Courriel"]
    }
    # Remap columns to target names (tolérant)
    rename = {}
    for k, lst in mapping.items():
        found = find_column(df, lst)
        if found:
            rename[found] = k
    df.rename(columns=rename, inplace=True)
    # Ne garde que les colonnes finales, dans l'ordre attendu
    final_cols = [k for k in mapping if k in df.columns]
    df = df[final_cols]

    for col in df.columns:
        if col != "Email":
            df[col] = df[col].astype(str).apply(format_name_advanced)
    df.drop_duplicates(subset=["Email"], inplace=True)
    df = df[~df["Ville"].str.lower().str.contains("emerainville|ozoir la ferriere", na=False)]

    ref_wb = openpyxl.load_workbook(BytesIO(ref_bytes))
    ref_ws = ref_wb.active
    col_widths = {c: ref_ws.column_dimensions[c].width for c in ref_ws.column_dimensions}

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Optin")
        wb = writer.book
        ws = wb["Optin"]
        fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True)
            cell.border = border
        for col_letter, width in col_widths.items():
            if width:
                ws.column_dimensions[col_letter].width = width
        for col in ws.iter_cols():
            if col[0].value == "Code Postal":
                for cell in col[1:]:
                    if isinstance(cell.value, str):
                        cleaned = cell.value.replace(" ", "").strip()
                        if cleaned.isdigit():
                            cell.value = int(cleaned)
                    cell.number_format = '0'
    out.seek(0)
    return out

# === GAG PROCESSING ===

EXCLUDED_DOMAINS = ["free.fr", "sfr.fr", "bouygtel.fr", "orange.fr", "bbox.fr", "laposte.net", "numericable.fr", "neuf.fr"]
KEYWORDS = ["concours", "jeu", "jeux"]

def is_excluded_email(email):
    email = str(email).lower()
    domain = email.split("@")[-1] if "@" in email else ""
    return (
        any(d in domain for d in EXCLUDED_DOMAINS) or
        any(k in email for k in KEYWORDS) or
        re.match(r".*\d{3,}.*", domain) or
        len(domain) > 20
    )

def tri_gagnants(df, nb_gagnants):
    civ_col = find_column(df, ["Civilite", "Civilité", "Civ", "Civilté"])
    if not civ_col:
        st.error("❌ Le fichier ne contient aucune colonne 'Civilité' (ou équivalent). Colonnes trouvées :\n" + str(list(df.columns)))
        return pd.DataFrame(), pd.DataFrame()
    df["__exclude__"] = df["Email"].apply(is_excluded_email)
    excl = df[df["__exclude__"]]
    main = df[~df["__exclude__"]]
    femmes = main[main[civ_col].astype(str).str.lower().str.startswith("femme")]
    hommes = main[main[civ_col].astype(str).str.lower().str.startswith("homme")]
    gagnants = pd.concat([femmes, hommes]).head(nb_gagnants)
    reservistes = pd.concat([main.drop(gagnants.index, errors="ignore"), excl])
    return gagnants.drop(columns="__exclude__"), reservistes.drop(columns="__exclude__")

def process_gag(df, nb_gagnants):
    # Colonnes minimales recherchées
    mapping = {
        "Civilite": ["Civilite", "Civilité", "Civ", "Civilté"],
        "Nom": ["Nom"],
        "Prenom": ["Prenom", "Prénom", "PrEnom"],
        "Adresse": ["Adresse"],
        "Code Postal": ["Code Postal", "CP"],
        "Ville": ["Ville"],
        "Pays": ["Pays"],
        "Tel": ["Tel", "Téléphone", "Numero de téléphone"],
        "Email": ["Email", "Mail", "Courriel"]
    }
    # Remap columns to target names (tolérant)
    rename = {}
    for k, lst in mapping.items():
        found = find_column(df, lst)
        if found:
            rename[found] = k
    df.rename(columns=rename, inplace=True)
    final_cols = [k for k in mapping if k in df.columns]
    df = df[final_cols]
    # Nettoyage
    df = df.applymap(remove_accents)
    for c in df.columns:
        if c.lower() != "email":
            df[c] = df[c].astype(str).apply(str.capitalize)
    if "Tel" in df.columns:
        df["Tel"] = df["Tel"].astype(str).apply(format_phone_number)
    df.fillna("", inplace=True)
    gagnants, reservistes = tri_gagnants(df, nb_gagnants)

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        gagnants.to_excel(writer, index=False, sheet_name="Gagnants")
        reservistes.to_excel(writer, index=False, sheet_name="Réservistes")
        wb = writer.book
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            adjust_column_width(ws)
    out.seek(0)
    return out

# === INTERFACE STREAMLIT ===

st.set_page_config(page_title="Traitement OPT & GAG", layout="centered")
st.title("📊 Traitement fichiers OPT / GAG")

traitement = st.selectbox("Traitement :", ["Opt-in partenaire (OPT)", "Tirages gagnants (GAG)"])
marque = st.radio("Marque :", ["FemmeActuelle.fr", "CuisineActuelle.fr"])
prefix = "FA.FR" if marque == "FemmeActuelle.fr" else "CA.FR"
nb_gagnants = st.number_input("🎯 Nombre de gagnants (GAG)", min_value=1, step=1, value=10) if traitement.startswith("Tirages") else None
uploaded_files = st.file_uploader("📂 Fichiers à traiter", type=["xls", "xlsx", "tsv"], accept_multiple_files=True)
go = st.button("🚀 Lancer le traitement")

def read_any_excel_or_tsv(raw_bytes, filename):
    import io
    ext = filename.lower().split('.')[-1]
    if ext == "xlsx":
        try:
            return pd.read_excel(io.BytesIO(raw_bytes), engine="openpyxl")
        except Exception as e:
            st.error(f"Erreur lecture Excel: {e}")
            return None
    elif ext == "xls":
        try:
            return pd.read_excel(io.BytesIO(raw_bytes), engine="xlrd")
        except Exception as e:
            st.error(f"Erreur lecture XLS : {e}")
            return None
    elif ext == "tsv":
        try:
            return pd.read_csv(io.BytesIO(raw_bytes), sep="\t", encoding="ISO-8859-1", dtype=str)
        except Exception as e:
            st.error(f"Erreur lecture TSV: {e}")
            return None
    else:
        st.error("Format non supporté")
        return None

if go and uploaded_files:
    zip_buf = BytesIO()
    zip_writer = ZipFile(zip_buf, "w")
    ref_opt = Path("ref_opt.xlsx").read_bytes() if Path("ref_opt.xlsx").exists() else None
    ref_gag = Path("ref_gag.xlsx").read_bytes() if Path("ref_gag.xlsx").exists() else None
    for f in uploaded_files:
        raw = f.read()
        df = read_any_excel_or_tsv(raw, f.name)
        if df is None:
            continue
        partner_match = re.search(r"(OPT|GAG)[ _-]*([A-Z0-9À-ÿ'’ ]+)", f.name, re.I)
        partenaire = partner_match.group(2).strip().upper() if partner_match else "PARTENAIRE"
        if traitement.startswith("Opt"):
            if not ref_opt:
                st.error("Fichier de référence OPT manquant !")
                continue
            output = process_opt(df, ref_opt)
            final_name = f"OPT {prefix} - {partenaire} GJ 2025.xlsx"
        else:
            if not ref_gag:
                st.error("Fichier de référence GAG manquant !")
                continue
            output = process_gag(df, nb_gagnants)
            final_name = f"GAG {prefix} - {partenaire} GJ 2025.xlsx"
        st.download_button(f"⬇️ Télécharger : {final_name}", output.getvalue(), file_name=final_name)
        zip_writer.writestr(final_name, output.getvalue())
    zip_writer.close()
    st.download_button("📦 Télécharger tout (ZIP)", zip_buf.getvalue(), file_name=f"Traitements_{traitement.split()[0]}.zip")
