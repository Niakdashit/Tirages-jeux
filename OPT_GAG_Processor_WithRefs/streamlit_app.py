
import streamlit as st
import pandas as pd
import openpyxl
import unicodedata
import re
from io import BytesIO
from zipfile import ZipFile, BadZipFile
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="OPT & GAG Processor", layout="centered")

st.title("📊 Traitement de fichiers OPT / GAG")
st.markdown("Nettoyez vos fichiers Excel/TSV pour FemmeActuelle.fr ou CuisineActuelle.fr, puis téléchargez-les individuellement ou en ZIP.")

# Sélection du traitement
treatment = st.selectbox("Quel traitement souhaitez-vous appliquer ?", ["Opt-in partenaire (OPT)", "Tirages gagnants (GAG)"])

# Sélection de la marque
brand = st.radio("Choisissez votre marque :", ["FemmeActuelle.fr", "CuisineActuelle.fr"])

brand_prefix = "FA.FR" if brand == "FemmeActuelle.fr" else "CA.FR"

# Nom du partenaire
partner_name = st.text_input("Nom du partenaire (ex: Homair)").strip()

# Upload des fichiers
uploaded_files = st.file_uploader("📂 Fichiers à traiter", type=["xls", "xlsx", "tsv"], accept_multiple_files=True)

# Fichier de référence requis pour les deux traitements
ref_file = None
ref_path = "references/ref_opt.xlsx" if treatment == "Opt-in partenaire (OPT)" else "references/ref_gag.xlsx"
ref_file = open(ref_path, "rb").read()

# Options de téléchargement
download_individual = st.checkbox("📥 Télécharger fichiers un par un", value=True)
download_zip = st.checkbox("📦 Télécharger tous les fichiers dans une archive ZIP")

# Fonctions partagées
def remove_accents_advanced(text):
    if isinstance(text, str):
        text = unicodedata.normalize('NFKD', text)
        text = re.sub(r'[\u0300-\u036f]', '', text)
        text = text.replace("Œ", "OE").replace("œ", "oe")
        return text.strip()
    return text

def format_name_advanced(value):
    if isinstance(value, str):
        value = remove_accents_advanced(value)
        return " ".join(word.capitalize() for word in value.split())
    return value

def convert_tsv_to_xlsx(input_bytes):
    df = pd.read_csv(BytesIO(input_bytes), sep="\t", encoding="ISO-8859-1", engine="python")
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    output.seek(0)
    return output.read()

# OPT Processor
def process_opt(input_bytes, ref_bytes, output_path):
    try:
        opt_xl = pd.ExcelFile(BytesIO(input_bytes))
        df = opt_xl.parse(opt_xl.sheet_names[0])
    except BadZipFile:
        st.error("❌ Fichier Excel OPT invalide.")
        return None

    col = next((c for c in df.columns if "partenaire -" in c.lower()), None)
    if not col:
        st.error("❌ Colonne 'Partenaire -' non trouvée.")
        return None
    df = df[df[col] == True]

    df.columns = [remove_accents_advanced(c) for c in df.columns]
    mapping = {
        "Civilite": ["Civilite", "CivilitE", "Civ"], "Nom": ["Nom"], "Prenom": ["Prenom", "PrEnom"],
        "Adresse": ["Adresse"], "Code Postal": ["Code Postal", "CP"], "Ville": ["Ville"],
        "Pays": ["Pays"], "Email": ["Email", "Mail", "Courriel"]
    }
    final_map = {n: k for k, v in mapping.items() for n in v if n in df.columns}
    df.rename(columns=final_map, inplace=True)
    df = df[list(final_map.values())]

    for c in df.columns:
        if c != "Email":
            df[c] = df[c].astype(str).apply(format_name_advanced)
    df.drop_duplicates(subset=["Email"], inplace=True)
    df = df[~df["Ville"].str.lower().str.contains("emerainville", na=False)]

    ref_wb = openpyxl.load_workbook(BytesIO(ref_bytes))
    ref_ws = ref_wb.active
    col_widths = {col: ref_ws.column_dimensions[col].width for col in ref_ws.column_dimensions}

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Optin", index=False)
        wb = writer.book
        ws = wb["Optin"]
        fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True)
            cell.border = border
        for col_letter, width in col_widths.items():
            if width:
                ws.column_dimensions[col_letter].width = width
        wb.save(output_path)
    return output_path

# GAG Processor
def process_gag(input_bytes, output_path):
    try:
        df = pd.read_excel(BytesIO(input_bytes), engine="openpyxl", dtype=str)
    except:
        df = pd.read_csv(BytesIO(input_bytes), sep="\t", encoding="utf-8", on_bad_lines="skip", dtype=str)

    df = df.shift(axis=1)
    df.rename(columns={"Merci de nous transmettre ici votre numéro de téléphone": "Tel"}, inplace=True)

    cols = ["Civilité", "Nom", "Prénom", "Adresse", "Complément d'adresse",
            "Code Postal", "Ville", "Pays", "Tel", "Email"]
    df = df[[c for c in cols if c in df.columns]]
    df.columns = [remove_accents_advanced(c) for c in df.columns]
    df = df.applymap(remove_accents_advanced)
    df.columns = [c.capitalize() if c.lower() != "email" else c for c in df.columns]
    for c in df.columns:
        if c.lower() != "email":
            df[c] = df[c].astype(str).apply(lambda x: x.capitalize())
    df.replace("Nan", "", inplace=True)
    df.fillna("", inplace=True)

    if "Tel" in df.columns:
        df["Tel"] = df["Tel"].astype(str).str.replace(".0", "", regex=False).str.zfill(10)
        df["Tel"] = df["Tel"].apply(lambda x: " ".join([x[i:i+2] for i in range(0, len(x), 2)]))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Gagnants", index=False)
        wb = writer.book
        ws = wb.active
        fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
        wb.create_sheet(title="Réservistes")
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
        wb.save(output_path)
    return output_path

# Traitement au clic
if uploaded_files and (treatment == "GAG - Traitement gagnants" or ref_file):
    zip_buffer = BytesIO()
    zip_archive = ZipFile(zip_buffer, "w")
    for file in uploaded_files:
        ext = file.name.split(".")[-1]
        in_bytes = file.read()
        if ext in ["tsv", "xls"]:
            in_bytes = convert_tsv_to_xlsx(in_bytes)
        out_name = f"OPT {brand_prefix} - {partner_name} GJ 2025 - {file.name}".replace(".tsv", ".xlsx").replace(".xls", ".xlsx")
        out_file = BytesIO()

        if treatment == "Opt-in partenaire (OPT)":
            path = process_opt(in_bytes, ref_file.read(), out_file)
        else:
            path = process_gag(in_bytes, out_file)

        if path:
            out_file.seek(0)
            if download_individual:
                st.download_button(f"⬇️ Télécharger {out_name}", out_file.getvalue(), file_name=out_name)
            if download_zip:
                zip_archive.writestr(out_name, out_file.getvalue())

    if download_zip:
        zip_archive.close()
        st.download_button("📦 Télécharger le ZIP complet", zip_buffer.getvalue(), file_name=f"OPT {brand_prefix} - GJ 2025.zip")
